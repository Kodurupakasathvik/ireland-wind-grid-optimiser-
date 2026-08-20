from pathlib import Path
from openpyxl import load_workbook


FILE = Path(
    "data/raw/System-Data-Qtr-Hourly-2026-V7.xlsx"
)


def main():
    if not FILE.exists():
        raise FileNotFoundError(
            f"Workbook not found: {FILE}"
        )

    print("=" * 90)
    print("EIRGRID WIND WORKBOOK INSPECTION")
    print("=" * 90)
    print(f"File: {FILE}")
    print()

    workbook = load_workbook(
        FILE,
        read_only=True,
        data_only=True,
    )

    print("SHEETS")
    print("-" * 90)

    for worksheet in workbook.worksheets:
        print(
            f"{worksheet.title!r} | "
            f"rows={worksheet.max_row} | "
            f"columns={worksheet.max_column}"
        )

    print()
    print("FIRST 12 ROWS OF EACH SHEET")
    print("-" * 90)

    for worksheet in workbook.worksheets:
        print()
        print(f"### SHEET: {worksheet.title}")

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(12, worksheet.max_row),
                values_only=True,
            ),
            start=1,
        ):
            values = list(row)

            print(
                f"{row_number:>3}: "
                + " | ".join(
                    repr(value)
                    for value in values[:20]
                )
            )

    workbook.close()


if __name__ == "__main__":
    main()